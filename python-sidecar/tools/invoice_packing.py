"""发票/箱单生成（FBA仓库版）

重构自原始脚本 `自动导出发票箱单_FBA仓库版.py`。
原脚本问题：硬编码本地文件路径 + 顶层执行无法被调用。
重构后：数据源从上传 bytes 读取，模板放 templates/，输出 bytes。

业务逻辑保持不变：
  数据源.xlsx（6 sheet）→ 按万邑通单号分组 → 套模板（德速/联宇）→ 输出 Excel

⚠️ 依赖模板文件（放到 tools/templates/ 下）：
  - 德速-模板.xlsx（德速渠道）
  - 联宇模板.xlsx（联宇渠道）
  - 商品尺寸申报清单.xlsx（商品申报单价/投保单价）
  这三个文件需你上传后放到 templates/ 目录。
"""

from __future__ import annotations

import io
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

# 模板目录（绝对路径，不受工作目录影响）
TEMPLATE_DIR = Path(__file__).parent / "templates"


def _load_data_source(file_bytes: bytes) -> dict[str, pd.DataFrame]:
    """读取数据源 Excel 的 6 个 sheet，返回 {sheet名: DataFrame}。"""
    xls = pd.ExcelFile(io.BytesIO(file_bytes))
    return {
        "装箱明细": xls.parse("装箱明细"),
        "地址库": xls.parse("地址库"),
        "箱规尺寸": xls.parse("箱规尺寸"),
        "万邑通重量表": xls.parse("万邑通重量表"),
        "FBA重量": xls.parse("FBA重量"),
        "新流程重量表": xls.parse("新流程重量表"),
    }


def _load_product_sizes(path: Optional[Path] = None) -> pd.DataFrame:
    """读取商品尺寸申报清单。"""
    path = path or (TEMPLATE_DIR / "商品尺寸申报清单.xlsx")
    if not path.exists():
        raise FileNotFoundError(f"商品尺寸申报清单未找到：{path}")
    return pd.ExcelFile(path).parse("Sheet1")


def _build_fba_mappings(address_df: pd.DataFrame) -> tuple[dict, dict]:
    """从地址库构建 FBA货件编号→追踪编号、→仓库代码 的映射。"""
    fba_mapping, warehouse_mapping = {}, {}
    for i in range(len(address_df)):
        cell = address_df.iloc[i, 0]
        if pd.isna(cell):
            continue
        cell_str = str(cell)
        if "货件编号：" not in cell_str:
            continue
        fba_number = cell_str.replace("货件编号：", "").strip()
        if i + 1 < len(address_df):
            tracking = address_df.iloc[i + 1, 0]
            if pd.notna(tracking) and "货件追踪编号:" in str(tracking):
                fba_mapping[fba_number] = str(tracking).replace("货件追踪编号:", "").strip()
        if i + 2 < len(address_df):
            warehouse = address_df.iloc[i + 2, 0]
            if pd.notna(warehouse) and "仓库代码：" in str(warehouse):
                warehouse_mapping[fba_number] = str(warehouse).replace("仓库代码：", "").strip()
    return fba_mapping, warehouse_mapping


def _get_weight_data(ref_no: str, data: dict[str, pd.DataFrame]) -> dict:
    """获取参考号对应的 {箱号: 重量}，按 WI/FBA/装箱明细/新流程 顺序查。"""
    weights = {}
    if ref_no.startswith("WI"):
        matching = data["万邑通重量表"][data["万邑通重量表"]["第三方入库单号"] == ref_no]
        for _, row in matching.iterrows():
            try:
                for box in range(int(row["起始箱号"]), int(row["终止箱号"]) + 1):
                    w = row["重量"]
                    if pd.notna(w) and box not in weights:
                        weights[box] = w
            except (ValueError, KeyError):
                continue
    elif ref_no.startswith("FBA"):
        matching = data["FBA重量"][data["FBA重量"]["万邑通单号"] == ref_no]
        for box_number, group in matching.groupby("数字箱号"):
            box_num = int(box_number) if isinstance(box_number, float) else box_number
            ws_in_group = group["重量"].dropna()
            if len(ws_in_group) > 0:
                weights[box_num] = ws_in_group.iloc[0]

    if not weights:
        matching = data["装箱明细"][data["装箱明细"]["万邑通单号"] == ref_no]
        for box_number, group in matching.groupby("数字箱号"):
            ws_in_group = group["重量"].dropna()
            if len(ws_in_group) > 0:
                weights[box_number] = ws_in_group.iloc[0]

    if not weights:
        matching = data["新流程重量表"][data["新流程重量表"]["万邑通单号"] == ref_no]
        for box_number, group in matching.groupby("数字箱号"):
            box_num = int(box_number) if isinstance(box_number, float) else box_number
            ws_in_group = group["重量"].dropna()
            if len(ws_in_group) > 0:
                weights[box_num] = ws_in_group.iloc[0]
    return weights


def _get_box_dims(package_name, box_size_df: pd.DataFrame):
    if pd.isna(package_name):
        return None, None, None
    m = box_size_df[box_size_df["规格"] == package_name]
    if not m.empty:
        row = m.iloc[0]
        return row["单箱长"], row["单箱宽"], row["单箱高"]
    return None, None, None


def _get_product_info(sku, product_sizes_df: pd.DataFrame):
    m = product_sizes_df[product_sizes_df["商品名称"] == sku]
    if not m.empty:
        row = m.iloc[0]
        return row["申报单价USD"], row["投保单价(RMB)"]
    sku_base = str(sku).split("*")[0].strip()
    m = product_sizes_df[product_sizes_df["商品名称"].str.contains(sku_base, na=False)]
    if not m.empty:
        row = m.iloc[0]
        return row["申报单价USD"], row["投保单价(RMB)"]
    return 100, 700


def _get_channel(ref_no, packing_df: pd.DataFrame) -> str:
    m = packing_df[packing_df["万邑通单号"] == ref_no]
    if not m.empty:
        ch = m["渠道"].iloc[0]
        if pd.notna(ch):
            s = str(ch)
            if "美森" in s: return "美国海速派-美森正班（包税）"
            if "合德" in s: return "美国海速派-合德以星快提（包税）"
            if "联宇" in s: return "联宇"
    return ""


def _get_product_names(ref_no, packing_df: pd.DataFrame):
    m = packing_df[packing_df["万邑通单号"] == ref_no]
    if not m.empty:
        pn = m["品名"].iloc[0]
        if pd.notna(pn) and "化纤" in str(pn):
            return "synthetic hair", "化纤制假发", "6704110000"
    return "wigs of human hair", "人发制假发", "6704200000"


def _set_cell_value(ws, row, col, value):
    """设置单元格值，处理合并单元格。"""
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            ws.cell(row=mr.min_row, column=mr.min_col, value=value)
            return
    ws.cell(row=row, column=col, value=value)


# 列映射（德速模板）
COL_MAPPING = {
    "Shipment ID": 1, "Reference ID": 2, "箱号段(起始-结束)": 3, "SKU": 4,
    "英文品名": 5, "中文品名": 6, "海关编码": 7, "Brand品牌": 8, "中文材质": 9,
    "用途": 10, "ASIN/销售链接": 11, "是否带电": 12, "型号": 13, "每箱数量": 14,
    "单位": 15, "每套个数": 16, "投保单价(RMB)": 17, "申报单价": 18, "图片": 19,
    "单箱重量(kg)": 20, "单箱长(cm)": 21, "单箱宽(cm)": 22, "单箱高(cm)": 23,
}


def generate_invoice_packing(data_source_bytes: bytes) -> dict[str, bytes]:
    """主入口：数据源 Excel bytes → {文件名: Excel bytes} 字典。

    Returns:
        {"导出_{单号}.xlsx": b"..."} 每个万邑通单号一个文件。
    """
    data = _load_data_source(data_source_bytes)
    product_sizes_df = _load_product_sizes()
    packing_df = data["装箱明细"]
    fba_mapping, warehouse_mapping = _build_fba_mappings(data["地址库"])
    tomorrow = (datetime.now() + timedelta(days=1)).strftime("%Y/%m/%d")

    results: dict[str, bytes] = {}
    for ref_no in packing_df["万邑通单号"].unique():
        ref_data = packing_df[packing_df["万邑通单号"] == ref_no]
        if ref_data.empty:
            continue

        total_boxes = ref_data["总箱数"].iloc[0]
        if ref_no.startswith("WI"):
            reference_id, warehouse_code = ref_no, "Winit-KY 41048"
        elif ref_no.startswith("FBA"):
            reference_id = fba_mapping.get(ref_no, ref_no)
            warehouse_code = warehouse_mapping.get(ref_no, "")
        else:
            reference_id, warehouse_code = ref_no, ""

        weight_data = _get_weight_data(ref_no, data)
        channel = _get_channel(ref_no, packing_df)
        english_name, chinese_name, customs_code = _get_product_names(ref_no, packing_df)
        is_lianyu = "联宇" in channel

        # 构建输出行
        output_rows = []
        for box_number, group in ref_data.groupby("数字箱号"):
            package_name = group["包装名称"].iloc[0]
            length, width, height = _get_box_dims(package_name, data["箱规尺寸"])
            if box_number in weight_data:
                weight = weight_data[box_number]
            elif length == 48.5 and width == 38 and height == 48.5:
                weight = 12
            else:
                weight = 15

            for _, row in group.iterrows():
                sku = row["产品名称"]
                platform = row["平台名称"]
                qty = int(row["计划数量"])
                declare_usd, insure_rmb = _get_product_info(sku, product_sizes_df)

                if str(customs_code).startswith("670420"):
                    cn_mat, en_mat = "真人发", "human hair"
                elif str(customs_code).startswith("670411"):
                    cn_mat, en_mat = "人造纤维", "synthetic fiber"
                elif "人发" in chinese_name:
                    cn_mat, en_mat = "真人发", "human hair"
                else:
                    cn_mat, en_mat = "化纤", "synthetic fiber"

                output_rows.append({
                    "Shipment ID": ref_no, "Reference ID": reference_id,
                    "箱号段(起始-结束)": box_number, "SKU": sku,
                    "英文品名": english_name, "中文品名": chinese_name,
                    "海关编码": customs_code, "Brand品牌": platform,
                    "中文材质": cn_mat, "英文材质": en_mat, "用途": "穿戴",
                    "ASIN/销售链接": "", "是否带电": "否", "型号": "无",
                    "每箱数量": qty, "单位": "个", "每套个数": "",
                    "投保单价(RMB)": insure_rmb if pd.notna(insure_rmb) else 700,
                    "申报单价": declare_usd if pd.notna(declare_usd) else 100,
                    "图片": "",
                    "单箱重量(kg)": weight if pd.notna(weight) else 15,
                    "单箱长(cm)": length if pd.notna(length) else "",
                    "单箱宽(cm)": width if pd.notna(width) else "",
                    "单箱高(cm)": height if pd.notna(height) else "",
                })

        # 套模板
        if is_lianyu:
            template_path = TEMPLATE_DIR / "联宇模板.xlsx"
            wb = load_workbook(template_path)
            ws = wb["Sheet1"]
            for r in range(ws.max_row, 1, -1):
                ws.delete_rows(r)
            used_boxes = set()
            for i, dr in enumerate(output_rows):
                tr = 2 + i
                while tr > ws.max_row:
                    ws.append([""] * 17)
                bn = dr["箱号段(起始-结束)"]
                if bn in used_boxes:
                    final_bn, box_count = 0, 0
                else:
                    final_bn, box_count = bn, 1
                    used_boxes.add(bn)
                ws.cell(tr, 1, ref_no)
                ws.cell(tr, 2, final_bn)
                ws.cell(tr, 3, fba_mapping.get(ref_no, ""))
                ws.cell(tr, 4, dr["SKU"])
                ws.cell(tr, 5, "")
                ws.cell(tr, 6, dr["英文品名"])
                ws.cell(tr, 7, dr["中文品名"])
                ws.cell(tr, 8, box_count)
                ws.cell(tr, 9, dr["申报单价"])
                ws.cell(tr, 10, dr["每箱数量"])
                ws.cell(tr, 11, dr["英文材质"])
                ws.cell(tr, 12, dr["中文材质"])
                ws.cell(tr, 13, dr["用途"])
                ws.cell(tr, 14, dr["海关编码"])
                ws.cell(tr, 15, dr["Brand品牌"])
                ws.cell(tr, 16, "")
                ws.cell(tr, 17, dr["中文材质"])
        else:
            template_path = TEMPLATE_DIR / "德速-模板.xlsx"
            wb = load_workbook(template_path)
            ws_inv = wb["发票填写模板"]
            _set_cell_value(ws_inv, 2, 2, ref_no)
            _set_cell_value(ws_inv, 3, 2, "FBA" if ref_no.startswith("FBA") else "热门海外仓")
            _set_cell_value(ws_inv, 3, 7, tomorrow)
            _set_cell_value(ws_inv, 4, 7, int(total_boxes))
            # 启运地
            m = packing_df[packing_df["万邑通单号"] == ref_no]
            wh = m["发货仓库"].iloc[0] if not m.empty else None
            loc = "青岛" if pd.notna(wh) and str(wh) in ["79", "139"] else "义乌"
            _set_cell_value(ws_inv, 4, 2, loc)
            _set_cell_value(ws_inv, 6, 2, warehouse_code)
            customs_status = "否" if (not m.empty and str(m["报关详情"].iloc[0]) == "不报关") else "是"
            _set_cell_value(ws_inv, 5, 7, customs_status)
            _set_cell_value(ws_inv, 7, 2, channel)
            for r in range(ws_inv.max_row, 20, -1):
                ws_inv.delete_rows(r)
            for i, dr in enumerate(output_rows):
                tr = 21 + i
                while tr > ws_inv.max_row:
                    ws_inv.append([""] * 23)
                for key, col in COL_MAPPING.items():
                    if key == "图片":
                        continue
                    _set_cell_value(ws_inv, tr, col, dr.get(key, ""))
            # 填写示例表
            ws_ex = wb["填写示例"]
            for r in range(ws_ex.max_row, 20, -1):
                ws_ex.delete_rows(r)
            for i, dr in enumerate(output_rows):
                tr = 21 + i
                while tr > ws_ex.max_row:
                    ws_ex.append([""] * 23)
                for key, col in COL_MAPPING.items():
                    if key == "图片":
                        continue
                    _set_cell_value(ws_ex, tr, col, dr.get(key, ""))

        buf = io.BytesIO()
        wb.save(buf)
        results[f"导出_{ref_no}.xlsx"] = buf.getvalue()

    return results
