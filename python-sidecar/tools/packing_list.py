"""箱单/发票生成工具

这是 MVP 的第一个工具。逻辑分三层：
  1. 数据模型（PackingList dataclass）—— 定义箱单包含哪些字段
  2. 从 Excel 读取订单数据 —— 现实中订单常以 Excel 形式给出
  3. 生成箱单 Excel 文件 —— 套用格式输出

⚠️ 当前用的是通用模板格式。要贴合你的实际业务，
   需要你提供一份真实的箱单/发票样例（Excel 或 PDF），
   我据此调整字段和模板格式。
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import date
from typing import List, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter


@dataclass
class PackingItem:
    """箱单中的一行商品明细"""
    description: str          # 品名
    quantity: int             # 数量
    cartons: int              # 箱数
    gross_weight: float       # 毛重 kg
    net_weight: float         # 净重 kg
    cbm: float                # 体积 m³
    unit: str = "PCS"         # 单位


@dataclass
class PackingList:
    """箱单主数据结构"""
    shipper: str              # 发货人
    consignee: str            # 收货人
    invoice_no: str           # 发票号
    invoice_date: date        # 日期
    port_of_loading: str      # 起运港
    port_of_destination: str  # 目的港
    shipping_marks: str = "N/M"  # 唛头，默认 No Marks
    items: List[PackingItem] = field(default_factory=list)

    @property
    def total_quantity(self) -> int:
        return sum(i.quantity for i in self.items)

    @property
    def total_cartons(self) -> int:
        return sum(i.cartons for i in self.items)

    @property
    def total_gross_weight(self) -> float:
        return round(sum(i.gross_weight for i in self.items), 2)

    @property
    def total_net_weight(self) -> float:
        return round(sum(i.net_weight for i in self.items), 2)

    @property
    def total_cbm(self) -> float:
        return round(sum(i.cbm for i in self.items), 3)


def parse_orders_excel(file_bytes: bytes) -> PackingList:
    """从订单 Excel 读取数据，构建 PackingList。

    ⚠️ 这里假设了一个通用的订单 Excel 格式：
       - 表头行，列含：shipper/consignee/invoice_no/date/pol/pod
         以及明细行：description/quantity/cartons/gw/nw/cbm
    实际格式要按你提供的样例调整。
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # 简化：假设第一张表第 1 行是抬头信息（key:value 两列），第 3 行起是明细表头
    header: dict[str, str] = {}
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        # 占位：真实解析逻辑待样例确认后补全
        pass

    # TODO: 待用户提供真实样例后实现精确解析
    return PackingList(
        shipper="（待解析）",
        consignee="（待解析）",
        invoice_no="PL-PLACEHOLDER",
        invoice_date=date.today(),
        port_of_loading="",
        port_of_destination="",
    )


def build_packing_list_excel(data: PackingList) -> bytes:
    """根据 PackingList 数据生成箱单 Excel 文件（bytes）。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Packing List"

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 列宽
    widths = [22, 30, 10, 10, 12, 12, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 标题
    ws.merge_cells("A1:G1")
    ws["A1"] = "PACKING LIST"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = center

    # 抬头信息
    info = [
        ("Shipper:", data.shipper, "Consignee:", data.consignee),
        ("Invoice No.:", data.invoice_no, "Date:", data.invoice_date.isoformat()),
        ("Port of Loading:", data.port_of_loading, "Port of Destination:", data.port_of_destination),
        ("Shipping Marks:", data.shipping_marks, "", ""),
    ]
    r = 3
    for label1, val1, label2, val2 in info:
        ws.cell(r, 1, label1).font = bold
        ws.cell(r, 2, val1).alignment = left
        if label2:
            ws.cell(r, 4, label2).font = bold
            ws.cell(r, 5, val2).alignment = left
        r += 1

    # 明细表头
    r += 1
    headers = ["No.", "Description", "Qty", "Cartons", "G.W.(kg)", "N.W.(kg)", "CBM"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(r, c, h)
        cell.font = bold
        cell.alignment = center
        cell.border = border

    # 明细行
    for idx, item in enumerate(data.items, 1):
        r += 1
        row = [idx, item.description, item.quantity, item.cartons,
               item.gross_weight, item.net_weight, item.cbm]
        for c, val in enumerate(row, 1):
            cell = ws.cell(r, c, val)
            cell.border = border
            cell.alignment = left if c == 2 else center

    # 合计行
    r += 1
    ws.cell(r, 1, "TOTAL").font = bold
    ws.cell(r, 3, data.total_quantity).font = bold
    ws.cell(r, 4, data.total_cartons).font = bold
    ws.cell(r, 5, data.total_gross_weight).font = bold
    ws.cell(r, 6, data.total_net_weight).font = bold
    ws.cell(r, 7, data.total_cbm).font = bold
    for c in range(1, 8):
        ws.cell(r, c).border = border
        ws.cell(r, c).alignment = center

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
