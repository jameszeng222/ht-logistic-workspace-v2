"""Excel/CSV 数据分析工具

输入：Excel(.xlsx) 或 CSV 文件的 bytes
输出：dict（JSON），包含每列的自动统计：
  - 数值列：count/mean/std/min/max/median/分布直方图
  - 分类列：unique count / top 频次 / 缺失数
  - 时间列：范围 / 跨度 / 趋势
  - 其它列：类型 + 样本 + 缺失数

设计目标：用户拖一个 Excel 进来，立刻得到一份"这份数据长什么样"的报告，
不用自己写 pandas 代码。是物流工作台里 AI 能力与工具能力的衔接点——
分析结果交给 AI 进一步解读、出方案、生成报告。
"""

from __future__ import annotations

import io
import json
from typing import Any

import pandas as pd
import numpy as np
from openpyxl import load_workbook


# 数值统计保留几位小数，避免浮点噪声
_ROUND = 4
# 直方图分桶数
_HIST_BINS = 10
# 分类列 top 频次显示几个
_TOP_N = 5
# 样本预览行数
_SAMPLE_N = 3


def analyze_excel_data(
    data: bytes,
    file_name: str = "upload",
    input_format: str = "auto",
    header_row: int = 1,
    sheet_name: str = "",
) -> dict[str, Any]:
    """分析 Excel/CSV 数据，返回结构化统计报告。

    Args:
        data: 文件 bytes（支持 .xlsx/.xls/.csv）
        file_name: 原始文件名（用于报告标题，可选）

    Returns:
        {
            "file": str,
            "shape": [rows, cols],
            "columns": [
                {
                    "name": str,
                    "dtype": str,           # pandas dtype 简化名
                    "kind": "numeric" | "categorical" | "datetime" | "text" | "mixed",
                    "missing": int,
                    "missing_pct": float,
                    "unique": int,
                    "stats": {...},         # 按 kind 不同
                    "sample": [...],
                },
                ...
            ],
            "correlations": [...],          # 数值列之间的相关系数（>0.3 才列）
            "summary": str,                 # 一句话摘要
        }
    """
    # 1. 读文件
    df = _read_data(data, file_name, input_format, header_row, sheet_name)

    # 2. 逐列分析
    columns = []
    numeric_cols: list[str] = []
    for col in df.columns:
        info = _analyze_column(df[col])
        columns.append(info)
        if info["kind"] == "numeric":
            numeric_cols.append(col)

    # 3. 数值列相关性（>0.3 才报告）
    correlations = []
    if len(numeric_cols) >= 2:
        corr = df[numeric_cols].corr(numeric_only=True)
        for i, a in enumerate(numeric_cols):
            for j, b in enumerate(numeric_cols):
                if j <= i:
                    continue
                v = corr.loc[a, b]
                if pd.isna(v):
                    continue
                v = round(float(v), _ROUND)
                if abs(v) >= 0.3:
                    correlations.append({"a": a, "b": b, "value": v})

    # 4. 一句话摘要
    summary = (
        f"{file_name}: {df.shape[0]} 行 × {df.shape[1]} 列，"
        f"其中 {len(numeric_cols)} 个数值列、"
        f"{sum(1 for c in columns if c['kind'] == 'categorical')} 个分类列、"
        f"{sum(1 for c in columns if c['kind'] == 'datetime')} 个时间列。"
    )

    return {
        "file": file_name,
        "shape": [int(df.shape[0]), int(df.shape[1])],
        "columns": columns,
        "correlations": correlations,
        "summary": summary,
    }


def export_data(
    data: bytes,
    file_name: str = "upload",
    input_format: str = "auto",
    header_row: int = 1,
    sheet_name: str = "",
    output_format: str = "xlsx",
    template_data: bytes | None = None,
) -> bytes:
    """按配置读取源数据，并导出 CSV 或套用用户 Excel 模板。"""
    df = _read_data(data, file_name, input_format, header_row, sheet_name)
    if output_format == "csv":
        return df.to_csv(index=False).encode("utf-8-sig")
    if output_format != "xlsx":
        raise ValueError(f"不支持的输出格式：{output_format}")

    if not template_data:
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="数据分析")
        return output.getvalue()

    workbook = load_workbook(io.BytesIO(template_data))
    worksheet = workbook.active
    header_candidates = []
    for row_index in range(1, min(worksheet.max_row, 20) + 1):
        values = [worksheet.cell(row_index, column).value for column in range(1, worksheet.max_column + 1)]
        count = sum(bool(value is not None and str(value).strip()) for value in values)
        if count:
            header_candidates.append((count, row_index, values))
    if not header_candidates:
        raise ValueError("输出模板中没有找到表头")
    _, template_header_row, header_values = max(header_candidates, key=lambda item: (item[0], -item[1]))
    source_columns = {str(column).strip(): column for column in df.columns}
    mappings = [
        (column_index, source_columns[str(header).strip()])
        for column_index, header in enumerate(header_values, start=1)
        if header is not None and str(header).strip() in source_columns
    ]
    if not mappings:
        raise ValueError("模板表头与输入文件字段没有匹配项，请检查字段名称")

    for row in worksheet.iter_rows(min_row=template_header_row + 1, max_row=worksheet.max_row):
        for cell in row:
            cell.value = None
    for offset, (_, source_row) in enumerate(df.iterrows(), start=1):
        target_row = template_header_row + offset
        for target_column, source_column in mappings:
            worksheet.cell(target_row, target_column).value = _excel_value(source_row[source_column])

    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


# ============ 内部 ============

def _read_data(
    data: bytes,
    file_name: str,
    input_format: str = "auto",
    header_row: int = 1,
    sheet_name: str = "",
) -> pd.DataFrame:
    """按用户指定格式、表头行和工作表读取源数据。"""
    name = (file_name or "").lower()
    normalized_format = input_format if input_format in {"auto", "excel", "csv"} else "auto"
    is_csv = normalized_format == "csv" or (normalized_format == "auto" and name.endswith(".csv"))
    pandas_header = max(1, int(header_row or 1)) - 1
    if is_csv:
        for encoding in ("utf-8-sig", "utf-8", "gbk"):
            try:
                return pd.read_csv(io.BytesIO(data), encoding=encoding, header=pandas_header, sep=None, engine="python")
            except UnicodeDecodeError:
                continue
        raise ValueError("无法识别 CSV 文件编码")
    selected_sheet: str | int = sheet_name.strip() if sheet_name and sheet_name.strip() else 0
    return pd.read_excel(io.BytesIO(data), sheet_name=selected_sheet, header=pandas_header)


def _excel_value(value: Any) -> Any:
    """把 pandas/numpy 值转换为 openpyxl 可写入的原生值。"""
    if pd.isna(value):
        return None
    if isinstance(value, np.generic):
        return value.item()
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    return value


def _analyze_column(s: pd.Series) -> dict[str, Any]:
    """分析单列，按 kind 分支。"""
    name = str(s.name)
    missing = int(s.isna().sum())
    missing_pct = round(missing / len(s) * 100, 2) if len(s) else 0.0
    unique = int(s.nunique(dropna=True))

    # 判断 kind：优先尝试转数值/时间，回退分类/文本
    kind, dtype, stats = _infer_kind_and_stats(s)

    return {
        "name": name,
        "dtype": dtype,
        "kind": kind,
        "missing": missing,
        "missing_pct": missing_pct,
        "unique": unique,
        "stats": stats,
        "sample": _sample_values(s),
    }


def _infer_kind_and_stats(s: pd.Series) -> tuple[str, str, dict[str, Any]]:
    """推断列类型并算统计量。

    推断顺序：
      1. 已是数值 → numeric
      2. 已是时间 → datetime
      3. 字符串能整体转数值 → numeric（强转后统计）
      4. 字符串能整体转时间 → datetime
      5. unique <= max(unique, 50) → categorical
      6. 否则 → text
    """
    # 1. 数值
    if pd.api.types.is_numeric_dtype(s):
        return "numeric", str(s.dtype), _numeric_stats(s)

    # 2. 时间
    if pd.api.types.is_datetime64_any_dtype(s):
        return "datetime", str(s.dtype), _datetime_stats(s)

    # 3. 字符串 → 尝试转数值
    if s.dtype == object or pd.api.types.is_string_dtype(s):
        coerced = pd.to_numeric(s, errors="coerce")
        # 至少 80% 转成功才算数值列（防止"123abc"这类被误判）
        if coerced.notna().sum() / max(len(s), 1) >= 0.8:
            return "numeric", "float64 (coerced)", _numeric_stats(coerced)

    # 4. 字符串 → 尝试转时间
    if s.dtype == object or pd.api.types.is_string_dtype(s):
        coerced_dt = pd.to_datetime(s, errors="coerce")
        if coerced_dt.notna().sum() / max(len(s), 1) >= 0.8:
            return "datetime", "datetime64 (coerced)", _datetime_stats(coerced_dt)

    # 5. 分类（unique 少）
    if s.nunique(dropna=True) <= 50:
        return "categorical", str(s.dtype), _categorical_stats(s)

    # 6. 文本
    return "text", str(s.dtype), {}


def _numeric_stats(s: pd.Series) -> dict[str, Any]:
    """数值列统计量 + 直方图。"""
    s = s.dropna().astype(float)
    if len(s) == 0:
        return {"count": 0}

    # 分位数
    def _q(p: float) -> float:
        return round(float(s.quantile(p)), _ROUND) if len(s) else None

    # 直方图（bin 边界 + count）
    try:
        counts, edges = np.histogram(s, bins=_HIST_BINS)
        hist = [
            {"bin": [round(float(edges[i]), _ROUND), round(float(edges[i + 1]), _ROUND)],
             "count": int(counts[i])}
            for i in range(len(counts))
        ]
    except Exception:
        hist = []

    return {
        "count": int(len(s)),
        "mean": round(float(s.mean()), _ROUND),
        "std": round(float(s.std()), _ROUND) if len(s) > 1 else 0.0,
        "min": round(float(s.min()), _ROUND),
        "max": round(float(s.max()), _ROUND),
        "median": _q(0.5),
        "q25": _q(0.25),
        "q75": _q(0.75),
        "histogram": hist,
    }


def _datetime_stats(s: pd.Series) -> dict[str, Any]:
    """时间列统计。"""
    s = s.dropna()
    if len(s) == 0:
        return {"count": 0}
    mn, mx = s.min(), s.max()
    span_days = (mx - mn).days if pd.notna(mn) and pd.notna(mx) else None
    return {
        "count": int(len(s)),
        "min": str(mn),
        "max": str(mx),
        "span_days": span_days,
    }


def _categorical_stats(s: pd.Series) -> dict[str, Any]:
    """分类列统计：top 频次 + 占比。"""
    vc = s.value_counts(dropna=True)
    top = [
        {"value": str(k), "count": int(v), "pct": round(v / len(s) * 100, 2)}
        for k, v in vc.head(_TOP_N).items()
    ]
    return {
        "count": int(len(s)),
        "unique": int(len(vc)),
        "top": top,
    }


def _sample_values(s: pd.Series) -> list[str]:
    """取前 N 个非空值作为样本预览。"""
    vals = s.dropna().head(_SAMPLE_N).tolist()
    return [str(v) for v in vals]
